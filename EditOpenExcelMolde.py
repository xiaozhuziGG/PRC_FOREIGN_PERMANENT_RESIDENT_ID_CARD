"""
# @Time     : 2026/6/15
# @Author   :
# @File     : EditOpenExcelMolde.py
# @Software : PRC_FOREIGN_PERMANENT_RESIDENT_ID_CARD
# @Purpose  : 编辑开户 Excel 模板的 Tkinter 页面，
             根据 YAML 配置动态生成证件类型选择控件，
             并自动填充证件信息到 Excel 对应单元格。
"""

from __future__ import annotations

import json
import os
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import openpyxl
import xlrd
from xlutils.copy import copy as xl_copy
import xlwt

import IdCardGenerator
import YamlParse
from IdCardGenerator import IDKind, IDNOGenerator
from FieldMapping import GENDER_TO_BOP_GENDER,ID_KIND_TO_BOP_ID_KIND
# 资源文件的绝对路径
from Nationality import BASE_DIR



# ── 证件类型名称 → 证件对象构造器 映射
ID_KIND_TO_CARD_CLASS: Dict[str, type[IDNOGenerator]] = {
    IDKind.ID_CARD.value: IdCardGenerator.TypeSFZ,
    IDKind.FOREIGN_PERMANENT_RESIDENT2023.value: IdCardGenerator.TypeYJZ,
    IDKind.FOREIGN_PERMANENT_RESIDENT2017.value: IdCardGenerator.TypeYJZ2017,
    IDKind.GAT_PERMANENT_RESIDENT.value: IdCardGenerator.TypeGATJZZ,
    IDKind.HKG_MAC_PERMIT.value: IdCardGenerator.TypeGATXZ,
    IDKind.CTN_PERMIT.value: IdCardGenerator.TypeTWTXZ,
    IDKind.BUSINESS_LICENSE.value: IdCardGenerator.TypeYYZZ,
}

# 证件类型只能为营业执照的关联人
RELATED_TYPE_ORAGN_ONLY = ('客户主体','资产管理人','资产托管人')


class EditOpenExcelApp:
    """编辑开户 Excel 模板的主应用程序。"""

    def __init__(self, master: tk.Tk | tk.Toplevel) -> None:
        """
        初始化主窗口。

        :param master: tkinter 根窗口实例
        """
        self.master: tk.Tk | tk.Toplevel = master
        self.master.title("编辑开户Excel模板")
        self.master.geometry("500x600")

        # ── 状态变量 ──
        self.account_type: tk.StringVar = tk.StringVar(value="")
        self.excel_path: tk.StringVar = tk.StringVar()
        self.yaml_config: YamlParse.YamlConfig | None = None
        # 映射：OrganrelatedInfo 索引 → 证件类型 StringVar
        self.id_kind_vars: Dict[int, tk.StringVar] = {}

        self._init_ui()

    # ━━━ 界面初始化 ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

    def _init_ui(self) -> None:
        """创建全部 UI 控件。"""
        # ── 第一行：开户类型 ──
        tk.Label(self.master, text="开户类型:").grid(
            row=0, column=0, sticky="e", padx=5, pady=5
        )
        self.cb_account_type: ttk.Combobox = ttk.Combobox(
            self.master,
            textvariable=self.account_type,
            values=["机构", "产品"],
            state="readonly",
        )
        self.cb_account_type.grid(row=0, column=1, sticky="w", padx=5, pady=5)
        self.cb_account_type.bind("<<ComboboxSelected>>", self._on_account_type_change)

        # ── 第二行：Excel 文件选择 ──
        tk.Label(self.master, text="开户信息模板\nExcel文件:").grid(
            row=1, column=0, sticky="e", padx=5, pady=5
        )
        self.entry_excel: tk.Entry = tk.Entry(
            self.master, textvariable=self.excel_path, width=60
        )
        self.entry_excel.grid(row=1, column=1, sticky="ew", padx=5, pady=5)
        tk.Button(self.master, text="浏览", command=self._browse_excel).grid(
            row=1, column=2, padx=5, pady=5
        )
        self.master.columnconfigure(1, weight=1)

        # ── 第三行：配置信息显示区域（带滚动条） ──
        self.config_canvas: tk.Canvas = tk.Canvas(self.master, borderwidth=0)
        self.scrollbar: ttk.Scrollbar = ttk.Scrollbar(
            self.master, orient="vertical", command=self.config_canvas.yview
        )
        self.config_canvas.configure(yscrollcommand=self.scrollbar.set)

        self.config_frame_inner: ttk.Frame = ttk.Frame(self.config_canvas)
        self.config_frame_inner.bind(
            "<Configure>",
            lambda e: self.config_canvas.configure(
                scrollregion=self.config_canvas.bbox("all")
            ),
        )
        self.config_canvas.create_window(
            (0, 0), window=self.config_frame_inner, anchor="nw"
        )

        # 鼠标滚轮支持：聚焦 canvas 时滚动
        def _on_mousewheel(event: tk.Event) -> None:
            self.config_canvas.yview_scroll(
                -int(event.delta / 120), "units"
            )

        self.config_canvas.bind("<Enter>", lambda _: self.config_canvas.focus_set())
        self.config_canvas.bind("<MouseWheel>", _on_mousewheel)
        # Linux 兼容
        self.config_canvas.bind("<Button-4>", lambda _: self.config_canvas.yview_scroll(-3, "units"))
        self.config_canvas.bind("<Button-5>", lambda _: self.config_canvas.yview_scroll(3, "units"))

        self.config_canvas.grid(
            row=2, column=0, columnspan=3, sticky="nsew", padx=5, pady=5
        )
        self.scrollbar.grid(row=2, column=3, sticky="ns")
        self.master.rowconfigure(2, weight=1)

        # ── 第四行：编辑按钮 ──
        self.btn_edit: tk.Button = tk.Button(
            self.master, text="编辑", command=self._edit_excel
        )
        self.btn_edit.grid(row=3, column=1, pady=10)

    # ━━━ 事件处理 ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

    def _on_account_type_change(self, event: tk.Event | None = None) -> None:
        """
        开户类型变更事件处理器。

        根据选择的类型读取对应的 YAML 配置文件，解析后更新配置显示区域。

        :param event: Combobox 选择事件（可为 None）
        """
        config: YamlParse.YamlConfig | None = self._load_yaml_config()
        if config is not None:
            self._rebuild_config_widgets()

    def _browse_excel(self) -> None:
        """弹出文件选择对话框，选择 Excel 文件（支持 .xls 和 .xlsx）。"""
        file_path: str | None = filedialog.askopenfilename(
            filetypes=[
                ("Excel 文件", "*.xlsx *.xls"),
                ("xlsx 文件", "*.xlsx"),
                ("xls 文件", "*.xls"),
                ("所有文件", "*.*"),
            ]
        )
        if file_path:
            self.excel_path.set(file_path)

    # ━━━ YAML 加载与解析 ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

    def _load_yaml_config(self) -> YamlParse.YamlConfig | None:
        """
        根据当前开户类型读取并解析 YAML 配置文件。

        :return: 解析后的 YamlConfig 对象，失败时返回 None
        """
        account_type: str = self.account_type.get()
        yaml_path: Path
        if account_type == "机构":
            yaml_path = Path(BASE_DIR) / "resource" / "organ_open.yaml"
        else:
            yaml_path = Path(BASE_DIR) / "resource" / "product_open.yaml"

        if not yaml_path.exists():
            messagebox.showerror("错误", f"配置文件不存在: {yaml_path}")
            return None

        try:
            data: dict = YamlParse.load_yaml_file(yaml_path)
            self.yaml_config = YamlParse.parse_yaml_config(data)
            return self.yaml_config
        except Exception as e:
            messagebox.showerror("错误", f"解析配置文件出错: {e}")
            return None

    # ━━━ 配置控件重建 ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

    def _rebuild_config_widgets(self) -> None:
        """
        清空并重新创建配置显示区域中的证件类型选择控件。
        每个包含 ``organrelated_type`` 的 ``OrganrelatedInfo`` 显示一行：
        label（类型名称）+ combobox（IDKind 枚举值）。
        """
        # 清除旧控件
        for widget in self.config_frame_inner.winfo_children():
            widget.destroy()
        self.id_kind_vars.clear()

        if self.yaml_config is None:
            return

        id_kind_values: Tuple[str, ...] = tuple(
            member.value for member in IDKind
        )
        current_row: int = 0

        for sheet in self.yaml_config.sheets:
            # Sheet 标题
            lbl_sheet: tk.Label = tk.Label(
                self.config_frame_inner,
                text=f"📋 {sheet.sheet_name}",
                font=("", 11, "bold"),
            )
            lbl_sheet.grid(
                row=current_row, column=0, columnspan=3, sticky="w",
                padx=5, pady=(10, 2),
            )
            current_row += 1

            for idx, org_info in enumerate(sheet.organrelated_list):
                if org_info.organrelated_type is None:
                    continue

                # 全局唯一索引（用 sheet 名称 + idx 组合）
                unique_key: str = f"{id(sheet)}_{idx}"

                lbl: tk.Label = tk.Label(
                    self.config_frame_inner, text=f"  {org_info.organrelated_type}:"
                )
                lbl.grid(
                    row=current_row, column=0, sticky="e", padx=5, pady=2,
                )
                if org_info.organrelated_type in RELATED_TYPE_ORAGN_ONLY:
                    selectable_id_kind = (IDKind.BUSINESS_LICENSE.value)
                    # 默认显示的证件类型
                    default_id_kind_var: tk.StringVar =  tk.StringVar(value=IDKind.BUSINESS_LICENSE.value)
                else:
                    selectable_id_kind = id_kind_values
                    default_id_kind_var: tk.StringVar =  tk.StringVar(value=IDKind.ID_CARD.value)

                cb: ttk.Combobox = ttk.Combobox(
                    self.config_frame_inner,
                    textvariable=default_id_kind_var,
                    values=selectable_id_kind,
                    state="readonly",
                    width=35,
                )
                cb.grid(row=current_row, column=1, sticky="w", padx=5, pady=2)

                self.id_kind_vars[unique_key] = default_id_kind_var
                current_row += 1

    # ━━━ 创建证件信息对象 ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

    @staticmethod
    def _create_id_card(id_kind_name: str) -> IDNOGenerator:
        """
        根据证件类型名称创建对应的证件信息对象。

        部分证件类型（港澳台居民居住证、港澳通行证）需要额外参数，
        此处使用枚举中的第一个有效值作为默认值。

        :param id_kind_name: 证件类型名称（如 '居民身份证', '2023版外国人永久居留证' 等）
        :return: 证件信息对象实例
        """
        card_class: type[IDNOGenerator] | None = ID_KIND_TO_CARD_CLASS.get(id_kind_name)
        if card_class is None:
            return IdCardGenerator.TypeSFZ()

        try:
            if card_class is IdCardGenerator.TypeGATJZZ:
                return card_class(IdCardGenerator.GATPermanentResident.HKG_PERMANENT_RESIDENT.value)
            if card_class is IdCardGenerator.TypeGATXZ:
                return card_class(IdCardGenerator.HkgMacPermit.HKG_PERMIT.value)
            return card_class()
        except Exception:
            return IdCardGenerator.TypeSFZ()

    # ━━━ Excel 单元格操作 ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

    @staticmethod
    def _find_cell_xlsx(
        ws: openpyxl.worksheet.worksheet.Worksheet,
        search_text: str,
    ) -> openpyxl.cell.cell.Cell | None:
        """
        在 ``.xlsx`` 格式工作表中查找第一个包含指定文本的单元格。

        :param ws: openpyxl 工作表对象
        :param search_text: 要查找的文本
        :return: 匹配的单元格对象，未找到时返回 None
        """
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is not None and str(cell.value) == str(search_text):
                    return cell
        return None

    @staticmethod
    def _find_cell_xls(
        sheet: xlrd.sheet.Sheet,
        search_text: str,
    ) -> tuple[int, int] | None:
        """
        在 ``.xls`` 格式工作表中查找第一个包含指定文本的单元格。

        :param sheet: xlrd 工作表对象
        :param search_text: 要查找的文本
        :return: ``(row_index, col_index)`` 元组，未找到时返回 None
        """
        for row_idx in range(sheet.nrows):
            for col_idx in range(sheet.ncols):
                cell_value = sheet.cell_value(row_idx, col_idx)
                if cell_value is not None and str(cell_value) == str(search_text):
                    return (row_idx, col_idx)
        return None

    @staticmethod
    def _get_card_attr(card_obj: IDNOGenerator, attr_name: str) -> Any:
        """
        获取证件信息对象的属性值。

        :param card_obj: 证件信息对象
        :param attr_name: 属性名
        :return: 属性值（字符串形式），属性不存在时返回空字符串
        """
        if not hasattr(card_obj, attr_name):
            return ""
        value: Any = getattr(card_obj, attr_name)
        if value is None:
            return ""
        # 日期对象转字符串
        if hasattr(value, "strftime"):
            value = value.strftime("%Y%m%d")
        # 电话号码 / 传真号码前拼接 +86-
        if attr_name in ("fax_number", "phone_number"):
            value = f"+86-{value}" if value else value
        # 将字符串证件类型转换为BOP系统的证件类型
        if attr_name == 'id_kind':
            value = ID_KIND_TO_BOP_ID_KIND.get(value)
        if attr_name == 'gender':
            value = GENDER_TO_BOP_GENDER.get(value)
        return value

    # ━━━ 编辑 Excel 主逻辑 ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

    def _edit_excel(self) -> None:
        """
        执行 Excel 编辑操作。

        根据文件扩展名严格分支为两个独立的处理路径：

        - ``.xlsx`` → 使用 ``openpyxl`` 读写
        - ``.xls``  → 使用 ``xlrd`` + ``xlutils`` + ``xlwt`` 读写

        两个路径的业务逻辑完全相同：
        1. 按 ``sheet_name`` 定位工作表；
        2. 遍历 ``OrganrelatedInfo``，通过 ``field_mapping`` 的 key 查找单元格；
        3. 编辑该单元格右侧的单元格，填充证件信息对象的对应属性值；
        4. 若 ``OrganrelatedInfo`` 含 ``organrelated_type``，则使用下拉选框指定的证件类型；
           否则默认为"居民身份证"。
        """
        # ── 校验 ──
        excel_path: str = self.excel_path.get().strip()
        if not excel_path:
            messagebox.showerror("错误", "请选择 Excel 文件")
            return
        if self.yaml_config is None:
            messagebox.showerror("错误", "请先选择开户类型以加载配置")
            return

        file_ext: str = os.path.splitext(excel_path)[1].lower()

        # ═════════════════════════════════════════════════════
        #  分支一：.xlsx 格式 — 使用 openpyxl
        # ═════════════════════════════════════════════════════
        if file_ext == ".xlsx":
            try:
                wb: openpyxl.Workbook = openpyxl.load_workbook(excel_path)
            except FileNotFoundError:
                messagebox.showerror("错误", f"文件不存在: {excel_path}")
                return
            except Exception as e:
                messagebox.showerror("错误", f"打开 xlsx 文件失败: {e}")
                return

            try:
                for sheet in self.yaml_config.sheets:
                    if sheet.sheet_name not in wb.sheetnames:
                        continue

                    ws: openpyxl.worksheet.worksheet.Worksheet = wb[sheet.sheet_name]

                    for idx, org_info in enumerate(sheet.organrelated_list):
                        # 确定证件类型
                        unique_key: str = f"{id(sheet)}_{idx}"
                        if (org_info.organrelated_type is not None
                                and unique_key in self.id_kind_vars):
                            id_kind_name: str = self.id_kind_vars[unique_key].get()
                        else:
                            id_kind_name = IDKind.ID_CARD.value

                        # 创建证件信息对象
                        card_obj: IDNOGenerator = self._create_id_card(id_kind_name)

                        # 遍历 field_mapping 查找并编辑单元格
                        for key, value in org_info.field_mapping.items():
                            attr_name: str = value
                            cell = self._find_cell_xlsx(ws, key)
                            if cell is not None:
                                # 目标单元格：当前单元格的右侧
                                target_cell = ws.cell(
                                    row=cell.row, column=cell.column + 1
                                )
                                try:
                                    attr_value: Any = self._get_card_attr(
                                        card_obj, attr_name
                                    )
                                    target_cell.value = attr_value
                                except Exception as e:
                                    print(f"报错信息为{e},查找{card_obj}中的{attr_name}")
                                    

                wb.save(excel_path)
                messagebox.showinfo("成功", "Excel 编辑完成！")

            except Exception as e:
                messagebox.showerror("错误", f"编辑 xlsx 出错: {e}")
            finally:
                wb.close()

        # ═════════════════════════════════════════════════════
        #  分支二：.xls 格式 — 使用 xlrd + xlutils + xlwt
        # ═════════════════════════════════════════════════════
        elif file_ext == ".xls":
            try:
                rb: xlrd.Book = xlrd.open_workbook(
                    excel_path, formatting_info=True
                )
            except FileNotFoundError:
                messagebox.showerror("错误", f"文件不存在: {excel_path}")
                return
            except Exception as e:
                messagebox.showerror("错误", f"打开 xls 文件失败: {e}")
                return

            try:
                # 将 xlrd 工作簿复制为可写的 xlwt 工作簿
                wb_xls: xlwt.Workbook = xl_copy(rb)
                xls_sheet_names: list[str] = rb.sheet_names()

                for sheet in self.yaml_config.sheets:
                    if sheet.sheet_name not in xls_sheet_names:
                        continue

                    # xlrd 只读 sheet
                    rd_sheet: xlrd.sheet.Sheet = rb.sheet_by_name(sheet.sheet_name)
                    # xlwt 可写 sheet（用序号定位）
                    sheet_index: int = xls_sheet_names.index(sheet.sheet_name)
                    wt_sheet: xlwt.Worksheet = wb_xls.get_sheet(sheet_index)

                    for idx, org_info in enumerate(sheet.organrelated_list):
                        # 确定证件类型
                        unique_key: str = f"{id(sheet)}_{idx}"
                        if (org_info.organrelated_type is not None
                                and unique_key in self.id_kind_vars):
                            id_kind_name: str = self.id_kind_vars[unique_key].get()
                        else:
                            id_kind_name = IDKind.ID_CARD.value

                        # 创建证件信息对象
                        card_obj = self._create_id_card(id_kind_name)

                        # 遍历 field_mapping 查找并编辑单元格
                        for key, value in org_info.field_mapping.items():
                            attr_name: str = value
                            cell_pos = self._find_cell_xls(rd_sheet, key)
                            if cell_pos is not None:
                                row_idx, col_idx = cell_pos
                                attr_value = self._get_card_attr(
                                    card_obj, attr_name
                                )
                                # 目标单元格：当前单元格的右侧
                                wt_sheet.write(row_idx, col_idx + 1, attr_value)

                wb_xls.save(excel_path)
                messagebox.showinfo("成功", "Excel 编辑完成！")

            except Exception as e:
                messagebox.showerror("错误", f"编辑 xls 出错: {e}")

        else:
            messagebox.showerror(
                "错误",
                f"不支持的文件格式: {file_ext}\n请选择 .xls 或 .xlsx 文件",
            )

    # ── 调试辅助：JSON 打印配置（保留） ─────────────────────

    def print_config_as_json(self) -> None:
        """
        将当前 YAML 配置以 JSON 格式打印到控制台（调试用）。
        """
        if self.yaml_config is None:
            print("{}")
            return
        result: dict = YamlParse.config_to_dict(self.yaml_config)
        print(json.dumps(result, ensure_ascii=False, indent=2))


def main(parent: tk.Misc | None = None) -> None:
    """启动编辑开户 Excel 模板程序。

    :param parent: 父窗口(tk.Tk 或 tk.Toplevel)，传 None 则创建独立窗口
    """
    if parent is not None:
        root: tk.Toplevel = tk.Toplevel(parent)
    else:
        root: tk.Tk = tk.Tk()
    app: EditOpenExcelApp = EditOpenExcelApp(root)
    root.mainloop()


if __name__ == "__main__":
    main()
